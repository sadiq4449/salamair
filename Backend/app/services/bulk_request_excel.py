"""Excel bulk preview and create for fare requests."""

from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.models.request import Request
from app.models.request_history import RequestHistory
from app.models.user import User
from app.services.request_codes import generate_request_code


MAX_BULK_ROWS = 500
EXPECTED_HEADERS = ("Route", "PAX", "Price", "Travel Date", "Return Date", "Notes")


def _parse_date(val: Any) -> tuple[date | None, str | None]:
    if val is None or val == "":
        return None, None
    if isinstance(val, datetime):
        return val.date(), None
    if isinstance(val, date):
        return val, None
    s = str(val).strip()
    if not s:
        return None, None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date(), None
        except ValueError:
            continue
    return None, "Invalid date format (use YYYY-MM-DD)"


def _validate_route(route: str) -> list[str]:
    errs: list[str] = []
    if not route or not str(route).strip():
        errs.append("Route is required")
        return errs
    r = str(route).strip()
    if len(r) > 100:
        errs.append("Route must be at most 100 characters")
    if not re.search(r"[A-Za-z]{2,}", r):
        errs.append("Invalid route format")
    return errs


def _row_dict(ws, row_idx: int) -> dict[str, Any]:
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    row = ws[row_idx]
    data: dict[str, Any] = {}
    for i, h in enumerate(headers):
        if i < len(row):
            data[h] = row[i].value
    return data


def parse_excel_rows(content: bytes) -> tuple[list[dict[str, Any]], str | None]:
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        return [], f"Invalid Excel file: {e}"
    ws = wb.active
    if ws.max_row is None or ws.max_row < 2:
        return [], "No data rows found"
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    missing = [h for h in EXPECTED_HEADERS if h not in headers]
    if missing:
        return [], f"Missing columns: {', '.join(missing)}. Expected: {', '.join(EXPECTED_HEADERS)}"
    max_row = min(ws.max_row, MAX_BULK_ROWS + 1)
    rows_out: list[dict[str, Any]] = []
    for ri in range(2, max_row + 1):
        d = _row_dict(ws, ri)
        if all(d.get(h) in (None, "") for h in EXPECTED_HEADERS):
            continue
        rows_out.append({"row": ri - 1, "data": d})
    wb.close()
    if len(rows_out) > MAX_BULK_ROWS:
        return [], f"Maximum {MAX_BULK_ROWS} data rows allowed"
    return rows_out, None


def preview_bulk(content: bytes) -> dict[str, Any]:
    raw, err = parse_excel_rows(content)
    if err:
        return {"total_rows": 0, "valid_rows": 0, "invalid_rows": 0, "preview": [], "error": err}

    preview: list[dict[str, Any]] = []
    valid = invalid = 0
    for item in raw:
        ri = item["row"]
        d = item["data"]
        route = d.get("Route")
        route_s = str(route).strip() if route is not None else ""
        pax_raw = d.get("PAX")
        price_raw = d.get("Price")
        td, te = _parse_date(d.get("Travel Date"))
        rd, rde = _parse_date(d.get("Return Date"))
        notes = d.get("Notes")
        notes_s = str(notes).strip()[:2000] if notes is not None else ""

        errors: list[str] = []
        errors.extend(_validate_route(route_s))

        pax = None
        try:
            pax = int(pax_raw) if pax_raw is not None and str(pax_raw).strip() != "" else None
        except (TypeError, ValueError):
            errors.append("PAX must be a whole number")
        if pax is not None and pax <= 0:
            errors.append("PAX must be greater than 0")

        price = None
        try:
            if price_raw is not None and str(price_raw).strip() != "":
                price = float(price_raw)
        except (TypeError, ValueError):
            errors.append("Price must be a number")
        if price is not None and price <= 0:
            errors.append("Price must be greater than 0")
        if price is not None and price > 99_999_999:
            errors.append("Price out of allowed range")

        if te:
            errors.append(te)
        if rde:
            errors.append(rde)
        if td and rd and rd < td:
            errors.append("Return date cannot be before travel date")

        ok = len(errors) == 0 and pax is not None and price is not None
        if ok:
            valid += 1
        else:
            invalid += 1

        preview.append(
            {
                "row": ri,
                "route": route_s,
                "pax": pax if pax is not None else 0,
                "price": int(price) if price is not None else 0,
                "valid": ok,
                "errors": errors if errors else None,
            }
        )

    return {
        "total_rows": len(preview),
        "valid_rows": valid,
        "invalid_rows": invalid,
        "preview": preview,
    }


def build_template_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Requests"
    ws.append(list(EXPECTED_HEADERS))
    ws.append(["KHI → DXB", 2, 50000, "2026-05-15", "2026-05-22", "Example"])
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def commit_bulk_upload(db: Session, acting_user: User, agent_owner: User, content: bytes) -> dict[str, Any]:
    raw, err = parse_excel_rows(content)
    if err:
        raise ValueError(err)

    results: list[dict[str, Any]] = []
    to_create: list[dict[str, Any]] = []

    for item in raw:
        ri = item["row"]
        d = item["data"]
        route_s = str(d.get("Route") or "").strip()
        pax_raw, price_raw = d.get("PAX"), d.get("Price")
        td, _ = _parse_date(d.get("Travel Date"))
        rd, _ = _parse_date(d.get("Return Date"))
        notes_s = str(d.get("Notes") or "").strip()[:2000] or None

        errors: list[str] = []
        errors.extend(_validate_route(route_s))
        try:
            pax = int(pax_raw)
        except (TypeError, ValueError):
            pax = 0
            errors.append("PAX must be a whole number")
        try:
            price = float(price_raw) if price_raw is not None and str(price_raw).strip() != "" else 0.0
        except (TypeError, ValueError):
            price = 0.0
            errors.append("Price must be a number")
        if pax <= 0:
            errors.append("PAX must be greater than 0")
        if price <= 0:
            errors.append("Price must be greater than 0")

        if errors:
            results.append({"row": ri, "status": "failed", "error": "; ".join(errors)})
            continue

        to_create.append(
            {
                "row": ri,
                "route": route_s,
                "pax": pax,
                "price": price,
                "travel_date": td,
                "return_date": rd,
                "notes": notes_s,
            }
        )

    created = 0
    from app.services.sla_service import sync_sla_for_request

    for row in to_create:
        code = generate_request_code(db)

        req = Request(
            request_code=code,
            agent_id=agent_owner.id,
            route=row["route"],
            pax=row["pax"],
            price=Decimal(str(row["price"])),
            travel_date=row["travel_date"],
            return_date=row["return_date"],
            notes=row["notes"],
            status="submitted",
            priority="normal",
        )
        db.add(req)
        db.flush()
        db.add(
            RequestHistory(
                request_id=req.id,
                action="bulk_uploaded",
                actor_id=acting_user.id,
                to_status="submitted",
                details=(
                    "Created via bulk Excel upload"
                    if acting_user.id == agent_owner.id
                    else f"Bulk Excel upload by admin for agent {agent_owner.name} ({agent_owner.email})"
                ),
            )
        )
        sync_sla_for_request(db, req)
        created += 1
        results.append({"row": row["row"], "request_code": req.request_code, "status": "created"})

    failed = sum(1 for r in results if r.get("status") == "failed")
    db.commit()

    return {
        "message": "Bulk upload completed",
        "total_rows": len(results),
        "created": created,
        "failed": failed,
        "results": sorted(results, key=lambda x: x["row"]),
    }
